import openpyxl
import pandas as pd
import utils.utils as utils
import utils.dbutils as dbutils
import matplotlib.pyplot as plt
import utils.sharepoint_wrapper as shwp
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime


def elaborate_calls_management_report():
    conn = dbutils.open_connection_with_scripting_account()
    query_for_studio = """WITH temp_ivr_calls AS(
                                    SELECT CONVERT(date, SWITCHOFFSET(started_at_utc, '-05:00')) AS date,
                                        interaction_id,
                                        type_of_call
                                        FROM jps.talkdesk_ivr_studio
                                        WHERE MONTH(CONVERT(date, SWITCHOFFSET(started_at_utc, '-05:00'))) = MONTH(GETDATE())
                                        AND type_of_call != 'Handled by agent')

                        SELECT date, interaction_id, type_of_call
                            FROM temp_ivr_calls
                            WHERE MONTH(date) = MONTH(GETDATE())
                            AND YEAR(date) = YEAR(GETDATE());"""
                
    query_for_contacts = """WITH temp_contacts_calls AS (
                                    SELECT CONVERT(date, SWITCHOFFSET(started_at_utc, '-05:00')) AS date,
                                        type_of_call =  CASE
                                                    WHEN contact_type = 'Answered' AND direction = 'IN' THEN 'Calls handled by agents'
                                                    WHEN contact_type = 'Connected' AND direction = 'OUT' THEN 'Outbound Calls'
                                                    WHEN (contact_type = 'Abandoned' OR contact_type = 'Short Abandoned') AND direction = 'IN' THEN 'Abandoned Calls'
                                                    WHEN contact_type = 'Missed' AND direction = 'IN' THEN 'Missed Calls'
                                                    ELSE 'NO APL'
                                                        END
                                        FROM jps.talkdesk_calls_contacts)

                        SELECT date, type_of_call
                            FROM temp_contacts_calls
                            WHERE type_of_call != 'NO APL'
                            AND MONTH(date) = MONTH(GETDATE())
                            AND YEAR(date) = YEAR(GETDATE());"""
    
    contacts = pd.read_sql(query_for_contacts, conn)
    studio = pd.read_sql(query_for_studio, conn)
    studio["type_of_call"].replace(to_replace={
        "Main Menu": "Self-service Calls (Main Menu)",
        "Profile only": "Self-service Calls (Profile only)",
        "Other IVR": "No profile service calls",
        "IVR Abandoned": "IVR Abandoned Calls"
    }, inplace=True)
    studio_to_count = studio.drop_duplicates(subset=["interaction_id"])[["date", "type_of_call"]]
    data_to_count = pd.concat([contacts, studio_to_count])
    
    pivot = pd.pivot_table(data_to_count,
                           index="type_of_call",
                           columns="date",
                           aggfunc=len,
                           margins=True,
                           margins_name="Total")
    pivot = pivot.rename_axis("Type of call", axis="index")
    custom_order = [
        "Calls handled by agents",
        "Outbound Calls",
        "Missed Calls",
        "Abandoned Calls",
        "IVR Abandoned Calls",
        "Self-service Calls (Main Menu)",
        "Self-service Calls (Profile only)",
        "No profile service calls",
        "Total"
    ]

    cat_type = pd.CategoricalDtype(categories=custom_order, ordered=True)
    pivot.index = pivot.index.astype(cat_type)
    pivot_sorted = pivot.sort_index()
    pivot_sorted = pivot_sorted.fillna(0)
    pivot_sorted = pivot_sorted.astype(int)
    print(pivot_sorted)

    pivot_sorted.to_excel("src\jps\Calls management report.xlsx", index=True)
    file_path = "src\jps\Calls management report.xlsx"
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    fill = PatternFill(fill_type="solid", fgColor="C0C0C0")
    font_header = Font(bold=True)
    alignment_header = Alignment(horizontal="center", vertical="center")
    for column_cells in sheet.columns:
        for cell in column_cells:
            cell.fill = fill
            cell.font = font_header
            cell.alignment = alignment_header

    for row_cells in sheet.iter_rows(min_row=2, max_row=9):
        for cell in row_cells:
            cell.fill = fill
            cell.font = font_header
            cell.alignment = alignment_header

    for row in sheet.iter_rows(min_row=2):
        for cell in row[1:]:
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(color="000000")
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            cell.border = border
            
    workbook.save("src\jps\Calls management report.xlsx")
    plt.figure(figsize=(15, 5))
    data_to_count['date'] = pd.to_datetime(data_to_count['date']).dt.strftime('%m-%d')
    data_to_count.groupby("date").type_of_call.value_counts().unstack(1).plot(kind="line")
    plt.xlabel('Date')  # Cambia la etiqueta del eje x
    plt.ylabel('Count')
    plt.title('Summary')
    legend = plt.legend(title='Call Type', loc='best', prop={'size': 8})
    legend.get_title().set_fontsize(10)
    plt.grid()

    # Save the line chart as an image
    chart_img_path = "line_chart.png"
    plt.tight_layout()
    plt.savefig(chart_img_path, bbox_inches='tight')
    plt.close()

    # Load the existing Excel file
    file_path = "src\jps\Calls management report.xlsx"
    book = load_workbook(filename=file_path)

    # Get the sheet where you want to add the chart
    sheet_name = "Sheet1"  # Replace with the name of the sheet where you want to add the chart
    sheet = book[sheet_name]

    # Insert the chart image in cells B12:K12 and B30:K30
    img = openpyxl.drawing.image.Image(chart_img_path)
    img.width = 600  # Adjust the width of the chart as per your preference
    img.height = img.width / 1.8

    sheet.add_image(img, "B12")  # The location where you want to add the chart in the sheet

    # Save the Excel file
    book.save(file_path)
    calls_management_folder_id = utils.get_config("dshub_sharepoint_config")["folders"]["jps_calls_management"]["id"]
    max_month = studio['date'].max().strftime('%B')
    file_name = f"Call Management Report {max_month}.xlsx"
    ctx = shwp.get_datascience_hub_ctx()
    file_path = "src\jps\Calls management report.xlsx"
    with open(file_path, "rb") as file:
        file_bytes = file.read()
        shwp.upload_file_to_sharepoint_folder(ctx, calls_management_folder_id, file_name, file_bytes)
                
                
def main(optional: list):
    """ Runs the elaborate_calls_management_report.
    Args:
        optional (int): Run mode.
    """

    match optional[0]:
        case 1:
            elaborate_calls_management_report()       