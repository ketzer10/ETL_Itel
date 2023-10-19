import openpyxl
import pandas as pd
import utils.utils as utils
import utils.dbutils as dbutils
import utils.sharepoint_wrapper as shwp
from datetime import datetime


subs_order_for_report = {
    "Interaction with agents": ["Bill & Balance", "Apply or Terminate Service",
                                "Outage report", "All other queries", "No option selected"],
    
    "No interaction with agents": ["Only virtual agent", "No agents available",
                                   "After business hours", "Holiday", "Error"]
}

def create_report_table(df):
    sorted_dates = sorted(df.date.unique())
    data_for_df = {
        "Date": sorted_dates
    }

    for date in sorted_dates:
        total_inters = 0
        for category in subs_order_for_report:
            total_row = 0
            subs = subs_order_for_report[category]
            for sub in subs:
                if sub not in data_for_df:
                    data_for_df[sub] = []
                count = len(df[(df.date == date) & (df.subcategory == sub)].interaction_id.unique())
                data_for_df[sub].append(count)
                total_row += count
            total_category_column_name = f"{category} Total"
            if total_category_column_name not in data_for_df:
                data_for_df[total_category_column_name] = []
            data_for_df[total_category_column_name].append(total_row)
            total_inters += total_row
            total_inters_column_name = "Total interactions"
        if total_inters_column_name not in data_for_df:
            data_for_df[total_inters_column_name] = []
        data_for_df[total_inters_column_name].append(total_inters)
    
    df_ = pd.DataFrame(data_for_df)
    columns_total = pd.DataFrame(["Total"] + df_.sum(numeric_only=True).values.tolist()).T
    columns_total.columns = df_.columns.copy()
    df_report = pd.concat([df_, columns_total], axis=0, ignore_index=True)
    print(df_report)
    
    # Create a new Excel file with formatting
    
    excel_file = 'src\jps\webchat_report.xlsx'
    writer = pd.ExcelWriter(excel_file, engine='openpyxl')
    df_report.to_excel(writer, index=False, sheet_name='Sheet1')

    # Get the workbook and worksheet objects
    workbook = writer.book
    worksheet = workbook['Sheet1']

    # Set the format for the grey cells in columns 6 and 12
    grey_fill = openpyxl.styles.PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    for col in [6, 12]:  # Columns 6 and 12 (index starts from 0)
        for cell in worksheet[openpyxl.utils.get_column_letter(col + 1)]:  # Column index starts from 1
            cell.fill = grey_fill

    # Set bold font for the values in the last row
    bold_font = openpyxl.styles.Font(bold=True)
    for col_num, value in enumerate(df_report.columns, start=1):
        cell = worksheet.cell(row=len(df_report)+1, column=col_num)
        cell.font = bold_font

    # Set bold font for the values in the last column
    for idx, value in enumerate(df_report['Total interactions'], start=1):
        cell = worksheet.cell(row=idx, column=len(df_report.columns)+1)
        cell.font = bold_font

    # Set bold font for the values in columns 6 and 12
    for col_num in [6, 12]:  # Columns 6 and 12 (index starts from 0)
        col_letter = openpyxl.utils.get_column_letter(col_num + 1)  # Column index starts from 1
        for idx, value in enumerate(df_report.iloc[:, col_num], start=1):  # Start from row 1 (header row)
            cell = worksheet.cell(row=idx+1, column=col_num+1)
            cell.font = bold_font

    # Set bold font for the values in the last row
    for col_num, value in enumerate(df_report.columns, start=1):
        cell = worksheet.cell(row=len(df_report)+2, column=col_num)
        cell.font = bold_font

    # Set bold font for the values in the last column
    for idx, value in enumerate(df_report['Total interactions'], start=1):
        cell = worksheet.cell(row=idx, column=len(df_report.columns))
        cell.font = bold_font

    # Set center alignment for all cells (excluding the last row and last column)
    center_alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    for row in worksheet.iter_rows(min_row=1, max_row=len(df_report)+1):
        for cell in row:
            cell.alignment = center_alignment

    # Save the file
    writer.save()

def elaborate_webchat_report():
    conn = dbutils.open_connection_with_scripting_account()
    query = """SELECT  CONVERT(date, SWITCHOFFSET(started_at_utc, '-05:00')) AS date,
       interaction_id,
       category,
       subcategory
    FROM jps.talkdesk_chat_studio
    WHERE MONTH(CONVERT(date, SWITCHOFFSET(started_at_utc, '-05:00'))) = MONTH(GETDATE())
    AND YEAR(CONVERT(date, SWITCHOFFSET(started_at_utc, '-05:00'))) = YEAR(GETDATE());"""
    
    df = pd.read_sql(query, conn)
    create_report_table(df)
    
    ctx = shwp.get_datascience_hub_ctx()
    webchat_folder_id = utils.get_config("dshub_sharepoint_config")["folders"]["jps_webchat"]["id"]
    file_path = "src\jps\webchat_report.xlsx"
    current = datetime.now()
    # month_name = current.strftime('%B')
    max_month = df['date'].max().strftime('%B')
    file_name = f"JPS WebChat Report {max_month}.xlsx"
    with open(file_path, "rb") as file:
        file_bytes = file.read()
        shwp.upload_file_to_sharepoint_folder(ctx, webchat_folder_id, file_name, file_bytes)

def main(optional: list):
    """Runs the elaborate_webchat_report

    Args:
        optional (list): _description_
    """
    
    match optional[0]:
        case 1:
            elaborate_webchat_report()